import time
from selenium import webdriver
import pyautogui
import re
import tkinter as tk
from tkinter import simpledialog
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import gspread
from oauth2client.service_account import ServiceAccountCredentials

def get_user_input(prompt):
    root = tk.Tk()
    root.withdraw()
    user_input = simpledialog.askstring(title="Dane użytkownika", prompt=prompt)
    return user_input

imie = get_user_input("Podaj imię:").upper().strip()
nazwisko = get_user_input("Podaj nazwisko:").upper().strip()
rok_rozliczenia = get_user_input("Podaj rok:").strip()

# Google Sheets authentication
SERVICE_ACCOUNT_FILE = 'excel.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SPREADSHEET_ID = '1k4UVgLa00Hqa7le3QPbwQMSXwpnYPlvcEQTxXqTEY4U'
SHEET_NAME_1 = 'ZP dane kont'
SHEET_NAME_2 = 'ZP status'

# Authenticate and initialize the Google Sheets client
credentials = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, SCOPES)
client = gspread.authorize(credentials)
ws1 = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME_1)
ws2 = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME_2)

driver = webdriver.Chrome()
url = 'https://www.formulare-bfinv.de/ffw/form/display.do?%24context=2802C5863D1DB0B5962F'
driver.get(url)

def kolejnyFormularz():
    time.sleep(0.1)
    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[1])
    time.sleep(0.2)
    driver.close()
    time.sleep(0.2)
    driver.switch_to.window(window_handles[0])
    time.sleep(0.2)
    new_url = 'https://www.formulare-bfinv.de/ffw/content.do'
    driver.get(new_url)

def wait_and_send_keys(xpath, value):
    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    element.send_keys(value)

def pierwsza_zgoda():
    driver.maximize_window()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="datenschutz"]/button'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/ul/li[3]/div/div/div[1]/a/div/div[2]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[5]/div/a/span'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[6]/div/a/span'))).click()

def clean_data(cell_value):
    if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
        return None
    if isinstance(cell_value, str):
        return re.sub(r'[.,\'"]', '', cell_value).strip()
    return cell_value

def PobieranieFormularza():
    time.sleep(0.1)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[3]/input'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[7]/div/div[1]/div/div[2]/p/a'))).click()
    time.sleep(1)
    pyautogui.hotkey('ctrl', 's')
    time.sleep(0.5)
    pyautogui.click(x=653, y=73)
    time.sleep(0.5)
    save_path = "C:\\Users\\Kamil\\Downloads"
    pyautogui.typewrite(save_path)
    time.sleep(0.5)
    pyautogui.press('enter')
    time.sleep(0.5)
    pyautogui.click(x=992, y=633)
    time.sleep(1)

data = ws1.get_all_values()
row = None
for i, row_data in enumerate(data):
    if row_data[0] == imie and row_data[1] == nazwisko:
        row = i + 1
        break

if row is None:
    print("Nie znaleziono danych dla podanej osoby.")
    driver.quit()
    exit()

status_data = ws2.get_all_values()
row2 = None
for i, row_data2 in enumerate(status_data):
    full_name = row_data2[0]
    if full_name:
        full_name_parts = full_name.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie  and row_data2[2] == rok_rozliczenia:
            row2 = i + 1
            break

imie = imie.upper().strip()
nazwisko = nazwisko.upper().strip()
biuro = data[row-1][2].upper().strip()
nr_telefonu = data[row-1][3]
email = data[row-1][4]
stan_cywilny = data[row-1][5].upper().strip()
numer_konta_bankowego = data[row-1][6]
swift = data[row-1][7].upper().strip()
finanzamt = data[row-1][8].upper()
steuernummer = data[row-1][9]
id_nr_meza = data[row-1][10]
id_nr_zony = data[row-1][11]
data_urodzenia_meza = data[row-1][12]
religia = data[row-1][13].upper().strip()
ulica = data[row-1][14].upper()
miejscowosc = data[row-1][15].upper()
data_slubu = data[row-1][16]
data_urodzenia_zony = data[row-1][17]
imie_zony = data[row-1][18]

imie_nazwisko_numer = clean_data(status_data[row2-1][0])
status_de = clean_data(status_data[row2-1][1]).upper()
rok_rozliczenia = clean_data(status_data[row2-1][2])
zwrot = clean_data(status_data[row2-1][3])
opiekun = clean_data(status_data[row2-1][4]).upper()
uwagi = clean_data(status_data[row2-1][5])
poinformowany = clean_data(status_data[row2-1][6]).upper().strip()
wyslany = clean_data(status_data[row2-1][7]).upper().strip()
fahrkosten = clean_data(status_data[row2-1][8])
ubernachtung = clean_data(status_data[row2-1][9])
h24 = clean_data(status_data[row2-1][10])
h8 = clean_data(status_data[row2-1][11])
w_kabinie = clean_data(status_data[row2-1][12])
an_und_ab = clean_data(status_data[row2-1][13])
kind = clean_data(status_data[row2-1][14])
cena = clean_data(status_data[row2-1][15]).strip()
status_platnosci = clean_data(status_data[row2-1][16]).upper().strip()
zaplacono = clean_data(status_data[row2-1][17])
forma_zaplaty = clean_data(status_data[row2-1][18]).upper().strip()
nr_faktury = clean_data(status_data[row2-1][19])
data_wystawienia_faktury = clean_data(status_data[row2-1][20])
ZarobkiMezaNiem = clean_data(status_data[row2-1][21])
ZarobkiZonyNiem = clean_data(status_data[row2-1][22])
nr22 = clean_data(status_data[row2-1][23])
nr23 = clean_data(status_data[row2-1][24])
nr25 = clean_data(status_data[row2-1][25])
nr26 = clean_data(status_data[row2-1][26])
nr27 = clean_data(status_data[row2-1][27])
pracodawca = clean_data(status_data[row2-1][28])
chorobowe = clean_data(status_data[row2-1][29])

klasa_pit1 = clean_data(status_data[row2-1][30])
if klasa_pit1:
    klasa_pit1 = int(klasa_pit1)
else:
    klasa_pit1 = 0
brutto_pit1 = clean_data(status_data[row2-1][31])
if brutto_pit1:
    brutto_pit1 = int(brutto_pit1)
else:
    brutto_pit1 = 0
podatek_pit1 = clean_data(status_data[row2-1][32])
if podatek_pit1:
    podatek_pit1 = int(podatek_pit1)
else:
    podatek_pit1 = 0
doplata_pit1 = clean_data(status_data[row2-1][33])
if doplata_pit1:
    doplata_pit1 = int(doplata_pit1)
else:
    doplata_pit1 = 0
koscielny_pit1 = clean_data(status_data[row2-1][34])
if koscielny_pit1:
    koscielny_pit1 = int(koscielny_pit1)
else:
    koscielny_pit1 = 0
kurzarbeitgeld_pit1 = clean_data(status_data[row2-1][35])
if kurzarbeitgeld_pit1:
    kurzarbeitgeld_pit1 = int(kurzarbeitgeld_pit1)
else:
    kurzarbeitgeld_pit1 = 0

klasa_pit2 = clean_data(status_data[row2-1][36])
if klasa_pit2:
    klasa_pit2 = int(klasa_pit2)
else:
    klasa_pit2 = 0
brutto_pit2 = clean_data(status_data[row2-1][37])
if brutto_pit2:
    brutto_pit2 = int(brutto_pit2)
else:
    brutto_pit2 = 0
podatek_pit2 = clean_data(status_data[row2-1][38])
if podatek_pit2:
    podatek_pit2 = int(podatek_pit2)
else:
    podatek_pit2 = 0
doplata_pit2 = clean_data(status_data[row2-1][39])
if doplata_pit2:
    doplata_pit2 = int(doplata_pit2)
else:
    doplata_pit2 = 0
koscielny_pit2 = clean_data(status_data[row2-1][40])
if koscielny_pit2:
    koscielny_pit2 = int(koscielny_pit2)
else:
    koscielny_pit2 = 0
kurzarbeitgeld_pit2 = clean_data(status_data[row2-1][41])
if kurzarbeitgeld_pit2:
    kurzarbeitgeld_pit2 = int(kurzarbeitgeld_pit2)
else:
    kurzarbeitgeld_pit2 = 0

klasa_pit3 = clean_data(status_data[row2-1][42])
if klasa_pit3:
    klasa_pit3 = int(klasa_pit3)
else:
    klasa_pit3 = 0
brutto_pit3 = clean_data(status_data[row2-1][43])
if brutto_pit3:
    brutto_pit3 = int(brutto_pit3)
else:
    brutto_pit3 = 0
podatek_pit3 = clean_data(status_data[row2-1][44])
if podatek_pit3:
    podatek_pit3 = int(podatek_pit3)
else:
    podatek_pit3 = 0
doplata_pit3 = clean_data(status_data[row2-1][45])
if doplata_pit3:
    doplata_pit3 = int(doplata_pit3)
else:
    doplata_pit3 = 0
koscielny_pit3 = clean_data(status_data[row2-1][46])
if koscielny_pit3:
    koscielny_pit3 = int(koscielny_pit3)
else:
    koscielny_pit3 = 0
kurzarbeitgeld_pit3 = clean_data(status_data[row2-1][47])
if kurzarbeitgeld_pit3:
    kurzarbeitgeld_pit3 = int(kurzarbeitgeld_pit3)
else:
    kurzarbeitgeld_pit3 = 0

def wypelnij_imie_nazwisko_st():
    driver.find_element(By.XPATH, '//*[@id="name"]').send_keys(nazwisko)
    driver.find_element(By.XPATH, '//*[@id="vorname"]').send_keys(imie)
    if steuernummer:
        driver.find_element(By.XPATH, '//*[@id="steuernummer"]').send_keys(steuernummer)

def parse_address(address):
    pattern = re.compile(r'(?P<street>[\w\sąćęłńóśźżĄĆĘŁŃÓŚŹŻ]+?)(?:\s+(?P<block>\d+))?(?:\s*(?P<part>[A-Z]))?(?:/(?P<apartment>\d+))?$')
    match = pattern.match(address.strip())
    if match:
        return match.groupdict()
    else:
        raise ValueError("Format adresu jest niepoprawny")


def przekształć_na_nazwisko_dla_żony(nazwisko_męża):
    if not nazwisko_męża:
        return ""

    nazwisko_męża = nazwisko_męża.strip().upper()

    if nazwisko_męża.endswith("SKI"):
        nazwisko_żony = nazwisko_męża[:-3] + "SKA"
    elif nazwisko_męża.endswith("CKI"):
        nazwisko_żony = nazwisko_męża[:-3] + "CKA"
    elif nazwisko_męża.endswith("DZKI"):
        nazwisko_żony = nazwisko_męża[:-4] + "DZKA"
    elif nazwisko_męża.endswith("DZKA"):
        nazwisko_żony = nazwisko_męża
    elif nazwisko_męża.endswith("CKI"):
        nazwisko_żony = nazwisko_męża[:-3] + "CKA"
    elif nazwisko_męża.endswith("DZKA"):
        nazwisko_żony = nazwisko_męża[:-4] + "DZKA"
    elif nazwisko_męża.endswith("EWSKI"):
        nazwisko_żony = nazwisko_męża[:-5] + "EWSKA"
    elif nazwisko_męża.endswith("EWICZ"):
        nazwisko_żony = nazwisko_męża[:-5] + "EWICZ"
    elif nazwisko_męża.endswith("ICZ"):
        nazwisko_żony = nazwisko_męża[:-2] + "ÓWNA"
    elif nazwisko_męża.endswith("Y"):
        nazwisko_żony = nazwisko_męża[:-1] + "A"
    else:
        nazwisko_żony = nazwisko_męża + "OWA"

    return nazwisko_żony.upper()
def wczytaj_dane_i_wypelnij():
    pierwsza_zgoda()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[2]/div[1]/a/span[1]/span'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[4]/div/div/div/div[4]/input'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="k1"]'))).click()
    time.sleep(2)

# WYPELNIANIE DANYCH MĘŻA 
    finanzamt = clean_data(data[row-1][8])
    if finanzamt:
        wait_and_send_keys('//*[@id="finanzamt"]', finanzamt.split(',')[0])

    if id_nr_meza:
        wait_and_send_keys('//*[@id="identifikationssnummer"]', id_nr_meza[:2])
        wait_and_send_keys('//*[@id="identifikationssnummer2"]', id_nr_meza[3:6])
        wait_and_send_keys('//*[@id="identifikationssnummer3"]', id_nr_meza[7:10])
        wait_and_send_keys('//*[@id="identifikationssnummer4"]', id_nr_meza[11:])
    
    wypelnij_imie_nazwisko_st()

    if data_urodzenia_meza:
        wait_and_send_keys('//*[@id="geburtsdatum"]', data_urodzenia_meza)
        pyautogui.press('enter')
        time.sleep(5)
    
    if religia:
        wait_and_send_keys('//*[@id="religion_barrierearm"]', religia)
    time.sleep(0.5)

    if ulica:
        address_components = parse_address(ulica)
        wait_and_send_keys('//*[@id="strasse_hausnummer"]', address_components['street'])
        if address_components['block']:
            wait_and_send_keys('//*[@id="hausnummer"]', address_components['block'])
        if address_components['apartment']:
            wait_and_send_keys('//*[@id="hausnummerzusatz"]', address_components['apartment'])
        if address_components['part']:
            wait_and_send_keys('//*[@id="ergaenzung"]', address_components['part'])
    else:
        print("Adres jest pusty dla podanej osoby.")
    
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="staat2"]'))).send_keys("POLEN")
    kod_pocztowy, miasto = miejscowosc.split(' ', 1)
    miasto = miasto.upper()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wohnort"]'))).send_keys(miasto)
    kod_pocztowy = kod_pocztowy.upper()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="postleitzahl4"]'))).send_keys(kod_pocztowy)

    if data_slubu:
        wait_and_send_keys('//*[@id="datum"]', data_slubu)
        driver.find_element(By.XPATH, '//*[@id="k9"]').click()
        pyautogui.press('enter')
        time.sleep(5)

    if data_slubu:
        nazwisko_zony = przekształć_na_nazwisko_dla_żony(nazwisko)
        wait_and_send_keys('//*[@id="abweichender_name"]', nazwisko_zony)
        wait_and_send_keys('//*[@id="vorname_ehefrau"]', imie_zony)

        driver.find_element(By.XPATH, '//*[@id="geburtsdatum2"]').send_keys(data_urodzenia_zony)
        pyautogui.press('enter')
        time.sleep(5)

        if id_nr_zony:
            wait_and_send_keys('//*[@id="identifikationssnummer5"]', id_nr_zony[:2])
            wait_and_send_keys('//*[@id="identifikationssnummer6"]', id_nr_zony[3:6])
            wait_and_send_keys('//*[@id="identifikationssnummer7"]', id_nr_zony[7:10])
            wait_and_send_keys('//*[@id="identifikationssnummer8"]', id_nr_zony[11:])
        wait_and_send_keys('//*[@id="religion2_barrierearm"]', religia)
        pyautogui.press('enter')

    Strona2 = driver.find_element("xpath","/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[10]").click()
    time.sleep(3)

    numer = clean_data(data[row-1][6])
    iban_parts = [numer[i:i+5] for i in range(0, len(numer), 5)]

    iban_fields = [
        '//*[@id="iban2"]', '//*[@id="iban8"]', '//*[@id="iban9"]',
        '//*[@id="iban10"]', '//*[@id="iban11"]', '//*[@id="iban13"]', '//*[@id="iban14"]'
    ]

    for iban_part, iban_field in zip(iban_parts, iban_fields):
        wait_and_send_keys(iban_field, iban_part)

    if swift:
        wait_and_send_keys('//*[@id="bic"]', swift)
    else:
        print("Swift jest pusty dla podanej osoby.")
    time.sleep(0.2)
    driver.find_element("xpath", '//*[@id="k10"]').click()
    time.sleep(0.2)


    if chorobowe:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="k_hinweis"]'))).click()
        time.sleep(1)
        wait_and_send_keys('//*[@id="betrag8"]', chorobowe)

    PobieranieFormularza()

if __name__ == "__main__":
    wczytaj_dane_i_wypelnij()

if koscielny_pit1 or koscielny_pit2 or koscielny_pit3:
    koscielny_pit1 = round(int(koscielny_pit1)) if koscielny_pit1 else 0
    koscielny_pit2 = round(int(koscielny_pit2)) if koscielny_pit2 else 0
    koscielny_pit3 = round(int(koscielny_pit3)) if koscielny_pit3 else 0

    kolejnyFormularz()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[5]/div[1]/a'))).click()
    wypelnij_imie_nazwisko_st()
    koscielny_Caly = koscielny_pit1 + koscielny_pit2 + koscielny_pit3
    wait_and_send_keys('//*[@id="eur"]', koscielny_Caly)
    PobieranieFormularza()
#-------------------------------------------------------------------------------------#
# 017 WA-EST
if 1==1:
    #Otwieranie strony WA-est
    kolejnyFormularz()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[7]/div[5]/div[2]/div[2]/div[2]/ul/li[9]/div[1]/a'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[7]/div/div[1]/div/div[4]/input'))).click()
    time.sleep(3)
    #Wypełnianie danych WA-est
    wypelnij_imie_nazwisko_st()

    if ZarobkiMezaNiem is None:
        ZarobkiMezaNiem = 0
    driver.find_element("xpath", '//*[@id="betrag3"]').send_keys(ZarobkiMezaNiem)
    time.sleep(.5)
    if ZarobkiZonyNiem is not None:
        driver.find_element("xpath", '//*[@id="betrag4"]').send_keys(ZarobkiZonyNiem)
        time.sleep(1)
    if ZarobkiZonyNiem is not None:
        ZarobkiMezaNiem = int(ZarobkiMezaNiem)
    if ZarobkiZonyNiem is not None:
        ZarobkiZonyNiem = int(ZarobkiZonyNiem)
        if ZarobkiMezaNiem + ZarobkiZonyNiem < 20000:
            driver.find_element("xpath", '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[35]/input').click()
        else:
            driver.find_element("xpath", '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[119]/input').click()
    else:
        if ZarobkiMezaNiem < 20000:
            driver.find_element("xpath", '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[35]/input').click()
        else:
            driver.find_element("xpath", '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[119]/input').click()
    PobieranieFormularza()
#-------------------------------------------------------------------------------------#
# 045 KIND 
def process_sentence(sentence):
    people = []
    # Podziel tekst według nowej linii
    lines = sentence.strip().split('\n')
    for line in lines:
        # Usuń białe znaki z początku i końca linii
        line = line.strip()
        if line:
            # Podziel każdą linię według spacji
            parts = line.split()
            if len(parts) >= 4:
                # Przypisz odpowiednie fragmenty do zmiennych
                name = parts[0]
                date = parts[1]
                status = ' '.join(parts[2:])
                # Dodaj dane osoby do listy
                people.append((name, date, status))
            else:
                print(f"Warning: line '{line}' does not contain enough parts")
    return people, len(people)

def remove_commas(text):
    return text.replace(",", "")

def wypelnij_formularz_dla_dziecka(index, person):
    kolejnyFormularz()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[17]/div[1]/a'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[4]/div/div/div/div[4]/input'))).click()
    time.sleep(3)
    wypelnij_imie_nazwisko_st()
    time.sleep(1)
    wait_and_send_keys('//*[@id="lfd_nr"]', str(index))
    if person[2] != 'BEZ KG':
        wait_and_send_keys('//*[@id="betrag"]', person[2])
    wait_and_send_keys('//*[@id="vorname2"]', person[0])
    wait_and_send_keys('//*[@id="fam_kasse"]', 'Sachsen')
    wait_and_send_keys('//*[@id="datum"]', person[1])
    pyautogui.press('enter')
    time.sleep(6)
    wait_and_send_keys('//*[@id="datum5"]', '0101')
    pyautogui.press('enter')
    time.sleep(5)
    wait_and_send_keys('//*[@id="datum6"]', '3112')
    pyautogui.press('enter')
    time.sleep(5)
    wait_and_send_keys('//*[@id="staat"]', 'POLEN')
    
    wait_and_send_keys('//*[@id="k1"]', '1')
    czyZonaty = clean_data(data[row-1][5])
    if czyZonaty == "Żonaty":
        wait_and_send_keys('//*[@id="k2"]', '1')
    PobieranieFormularza()

if kind is not None:
    kind = remove_commas(kind)
    people_data, number_of_people = process_sentence(kind)
    print(people_data)
    print(f"Number of children: {number_of_people}")  # Debugging
    
    for i in range(number_of_people):
        print("number_of_people" + str(i))
    for i, person in enumerate(people_data):
        print(f"Filling form for child {i + 1}: {person}")  # Debugging
        if i >= 5:
            break
        wypelnij_formularz_dla_dziecka(i + 1, person)

#-------------------------------------------------------------------------------------#
# 055 Anlage N
if 1==1:
    sum_brutto_1_5 = 0
    sum_podatek_1_5 = 0
    sum_doplata_1_5 = 0
    sum_koscielny_1_5 = 0
    sum_kurzarbeitgeld_1_5 = 0

    sum_brutto_6 = 0
    sum_podatek_6 = 0
    sum_doplata_6 = 0
    sum_koscielny_6 = 0
    sum_kurzarbeitgeld_6 = 0
    

    # Sumowanie wartości w ramach tej samej grupy klas
    if klasa_pit1 in range(1, 6):
        sum_brutto_1_5 += brutto_pit1
        sum_podatek_1_5 += podatek_pit1
        sum_doplata_1_5 += doplata_pit1
        sum_koscielny_1_5 += koscielny_pit1
        sum_kurzarbeitgeld_1_5 += kurzarbeitgeld_pit1
    elif klasa_pit1 == 6:
        sum_brutto_6 += brutto_pit1
        sum_podatek_6 += podatek_pit1
        sum_doplata_6 += doplata_pit1
        sum_koscielny_6 += koscielny_pit1
        sum_kurzarbeitgeld_6 += kurzarbeitgeld_pit1

    if klasa_pit2 in range(1, 6):
        sum_brutto_1_5 += brutto_pit2
        sum_podatek_1_5 += podatek_pit2
        sum_doplata_1_5 += doplata_pit2
        sum_koscielny_1_5 += koscielny_pit2
        sum_kurzarbeitgeld_1_5 += kurzarbeitgeld_pit2
    elif klasa_pit2 == 6:
        sum_brutto_6 += brutto_pit2
        sum_podatek_6 += podatek_pit2
        sum_doplata_6 += doplata_pit2
        sum_koscielny_6 += koscielny_pit2
        sum_kurzarbeitgeld_6 += kurzarbeitgeld_pit2

    if klasa_pit3 in range(1, 6):
        sum_brutto_1_5 += brutto_pit3
        sum_podatek_1_5 += podatek_pit3
        sum_doplata_1_5 += doplata_pit3
        sum_koscielny_1_5 += koscielny_pit3
        sum_kurzarbeitgeld_1_5 += kurzarbeitgeld_pit3
    elif klasa_pit3 == 6:
        sum_brutto_6 += brutto_pit3
        sum_podatek_6 += podatek_pit3
        sum_doplata_6 += doplata_pit3
        sum_koscielny_6 += koscielny_pit3
        sum_kurzarbeitgeld_6 += kurzarbeitgeld_pit3

    print("Sumy wartości dla klas 1-5:")
    print(f"Suma brutto: {sum_brutto_1_5}")
    print(f"Suma podatek: {sum_podatek_1_5}")
    print(f"Suma dopłata: {sum_doplata_1_5}")
    print(f"Suma kościelny: {sum_koscielny_1_5}")
    print(f"Suma kurzarbeitgeld: {sum_kurzarbeitgeld_1_5}")

    print("Sumy wartości dla klasy 6:")
    print(f"Suma brutto: {sum_brutto_6}")
    print(f"Suma podatek: {sum_podatek_6}")
    print(f"Suma dopłata: {sum_doplata_6}")
    print(f"Suma kościelny: {sum_koscielny_6}")
    print(f"Suma kurzarbeitgeld: {sum_kurzarbeitgeld_6}")

    if klasa_pit1 is not None:
        kolejnyFormularz()
        # Otwieranie strony Anlage N
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[7]/div[5]/div[2]/div[2]/div[2]/ul/li[21]/div[1]/a'))).click()
        time.sleep(5)
        # Zgody
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[4]/div/div/div/div[4]/input'))).click()
        time.sleep(5)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[1]/input'))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[4]/input'))).click()
        time.sleep(5)

        # Wypełnianie danych osobowych
        driver.find_element("xpath", '//*[@id="name"]').send_keys(nazwisko)
        driver.find_element("xpath", '//*[@id="name2"]').send_keys(imie)
        if steuernummer is not None:
            driver.find_element("xpath", '//*[@id="steuernummer"]').send_keys(steuernummer)
        time.sleep(1)

        # Wypełnianie danych ANLAGE N
        if klasa_pit1 in range(1, 6) or klasa_pit2 in range(1, 6) or klasa_pit3 in range(1, 6):
            values = [x for x in [klasa_pit1, klasa_pit2, klasa_pit3] if x > 0]
            if values:
                min_value = min(values)
            driver.find_element("xpath", '//*[@id="steuerklasse"]').send_keys(min_value)
            pyautogui.press('enter')
            if sum_brutto_1_5 is not None:
                driver.find_element("xpath", '//*[@id="betrag"]').send_keys(sum_brutto_1_5)
            if sum_podatek_1_5 is not None:
                driver.find_element("xpath", '//*[@id="betrag2"]').send_keys(sum_podatek_1_5)
            if sum_doplata_1_5 is not None:
                driver.find_element("xpath", '//*[@id="betrag3"]').send_keys(sum_doplata_1_5)
            if sum_koscielny_1_5 is not None:
                driver.find_element("xpath", '//*[@id="betrag4"]').send_keys(sum_koscielny_1_5)
        if klasa_pit1 == 6 or klasa_pit2 == 6 or klasa_pit3 == 6:
            if sum_brutto_6 is not None:
                driver.find_element("xpath", '//*[@id="betrag6"]').send_keys(sum_brutto_6)
            if sum_podatek_6 is not None:
                driver.find_element("xpath", '//*[@id="betrag7"]').send_keys(sum_podatek_6)
            if sum_doplata_6 is not None:
                driver.find_element("xpath", '//*[@id="betrag8"]').send_keys(sum_doplata_6)
            if sum_koscielny_6 is not None:
                driver.find_element("xpath", '//*[@id="betrag9"]').send_keys(sum_koscielny_6)

        if sum_kurzarbeitgeld_1_5 is not None or sum_kurzarbeitgeld_6 is not None:
            sum_kurzarbeitgeld = sum_kurzarbeitgeld_6 + sum_kurzarbeitgeld_1_5
            driver.find_element("xpath", '//*[@id="betrag30"]').send_keys(sum_kurzarbeitgeld)
        
        #Przejście na stronę 3
        time.sleep(2)
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[11]/input'))).click()

        time.sleep(4)
        wait_and_send_keys('//*[@id="anwendungen3"]','BERUFSKLEIDUNG')
        wait_and_send_keys('//*[@id="betrag39"]','110')

        time.sleep(.5)
        #Przejście na stronę 4
        driver.find_element("xpath",'/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[12]/input').click()
        time.sleep(3)

        if fahrkosten is not None:
            driver.find_element("xpath",'//*[@id="fahrkosten"]').send_keys('Fahrtkosten')
            driver.find_element("xpath",'//*[@id="betrag45"]').send_keys(fahrkosten)

        if ubernachtung is not None:
            driver.find_element("xpath",'//*[@id="uebernachtungskosten"]').send_keys('Ubernachtungskosten')
            driver.find_element("xpath",'//*[@id="betrag71"]').send_keys(ubernachtung)

        if h24 is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage3"]').send_keys(h24)

        if h8 is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage"]').send_keys(h8)

        if w_kabinie is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage4"]').send_keys(w_kabinie)

        if an_und_ab is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage2"]').send_keys(an_und_ab)
        if pracodawca is not None:
            driver.find_element("xpath",'//*[@id="betrag47"]').send_keys(pracodawca)
        time.sleep(1)

        #Pobieranie
        PobieranieFormularza()

#-------------------------------------------------------------------------------------#
#Vorsorgeaufwendungen
kolejnyFormularz()
#Otwieranie strony Vorsorgeaufwendungen
vorsorgeaufwendungen= driver.find_element("xpath", '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[35]/div[1]/a').click()
time.sleep(3)

#Zgoda nr1
driver.find_element("xpath",'/html/body/div[4]/form/div[4]/div/div/div/div[4]/input').click()
time.sleep(2)
wypelnij_imie_nazwisko_st()
#Zgoda nr2
driver.find_element("xpath",'//*[@id="k_hinweis"]').click()
time.sleep(3)

driver.find_element("xpath",'//*[@id="betrag"]').send_keys(nr23)
driver.find_element("xpath",'//*[@id="betrag5"]').send_keys(nr22)
driver.find_element("xpath",'//*[@id="betrag15"]').send_keys(nr25)
driver.find_element("xpath",'//*[@id="betrag18"]').send_keys(nr26)
if nr27 is not None:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_buttonPanel"]/div[1]/div/div[11]'))).click()
    wait_and_send_keys('//*[@id="betrag71"]', nr27)
time.sleep(.5)

#Pobieranie
PobieranieFormularza()
time.sleep(100)
import time
from selenium import webdriver
import pyautogui
from openpyxl import load_workbook
import tkinter as tk
from tkinter import simpledialog
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import re

# Funkcja do pobierania danych z okienka dialogowego
def get_user_input(prompt):
    root = tk.Tk()
    root.withdraw()  # Ukryj główne okno
    user_input = simpledialog.askstring(title="Dane użytkownika", prompt=prompt)
    return user_input

# Załadowanie EXCELA
wb = load_workbook('Dane1.xlsx')
ws = wb['ZP Status DE']


row = None
for i in range(1, ws.max_row + 1):
    full_name_cell = ws[f'A{i}'].value
    if full_name_cell is not None:
        full_name_parts = full_name_cell.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
            row = i
            break

# Włączenie strony do rozliczania (Logowanie)
driver = webdriver.Chrome()
url = 'https://www.elster.de/eportal/login/softpse'
driver.get(url)
driver.maximize_window()
# Definicja funkcji do parsowania adresu
def parse_address(address):
    pattern = re.compile(r'(?P<street>[\w\sąćęłńóśźżĄĆĘŁŃÓŚŹŻ]+)\s+(?P<block>\d+)(?:/(?P<apartment>\d+)(?P<part>[AB]?))?')
    match = pattern.match(address.strip())
    if match:
        return match.groupdict()
    else:
        raise ValueError("Format adresu jest niepoprawny")

# Wybór metody logowania
driver.find_element(By.XPATH, '//*[@id="submitButton_loginBox.file_cert"]').click()
time.sleep(.5)
pyautogui.click(x=115, y=488)
pyautogui.click(x=412, y=195)
pyautogui.press('enter')

# Wprowadzenie hasła i zatwierdzenie
driver.find_element(By.XPATH, '//*[@id="password"]').send_keys('GrandaKamil11071989!')
driver.find_element(By.XPATH, '//*[@id="bestaetigenButton"]').click()

try:
    element = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="temporaereaufgaben_nein_button"]'))
    )
    element.click()


except TimeoutException:
    # Element nie pojawił się w przeciągu 10 sekund, więc przechodzimy dalej
    pass
# Oczekiwanie na załadowanie strony i przejście do formularzy

meineFormulareURL= 'https://www.elster.de/eportal/meineformulare'
driver.get(meineFormulareURL)
# Pobranie roku rozliczenia i przejście do odpowiedniego formularza
rokRoliczenia = ws[f'C{row}'].value
czyZonaty = ws[f'V{row}'].value
if rokRoliczenia == 2023:
    if czyZonaty is not None:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="oeffneEntwurf_452354774"]'))).click()
    else:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="oeffneEntwurf_448520188"]'))).click()
elif rokRoliczenia == 2022:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="oeffneEntwurf_425366900"]'))).click()
elif rokRoliczenia == 2021:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="oeffneEntwurf_433741909"]'))).click()


def wait_and_send_keys(xpath, value):
    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    element.send_keys(value)

def zapiszKlase1():
    try:
        WebDriverWait(driver, .1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[0]"]'))).click()
    except:
        pass
    try:
        WebDriverWait(driver, .1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[1]"]'))).click()
    except:
        pass
    try:
        WebDriverWait(driver, .1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[2]"]'))).click()
    except:
        pass

def zapiszKlase6():
    try:
        WebDriverWait(driver, .1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[0]"]'))).click()
    except:
        pass
    try:
        WebDriverWait(driver, .1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[1]"]'))).click()
    except:
        pass


def klasa1_pr1():
    try:
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_fields(eruNArbLLStB_1_5_SumE0200002)"]', klasa)
        pyautogui.press('enter')
        if brutto is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200204)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200304)"]', podatek)
        if doplata is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200404)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200504)"]', koscielny)
        zapiszKlase1()
    except:
        pass
def klasa1_pr2():
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="AddMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf"]'))).click()
    try:
        if brutto is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200204)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200304)"]', podatek)
        if doplata is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200404)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200504)"]', koscielny)
        zapiszKlase1()
    except:
        pass
def klasa1_pr3():
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[1]"]'))).click()
    WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="AddMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf"]'))).click()
    try:
        if brutto is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200204)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200304)"]', podatek)
        if doplata is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200404)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200504)"]', koscielny)
        zapiszKlase1()
    except:
        pass
def klasa6_pr1():
    try:
        if brutto is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200202)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200302)"]', podatek)
        if doplata is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200402)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200502)"]', koscielny)
        zapiszKlase6()
    except:
        pass
def klasa6_pr2():
    try:
        WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[5]/main/div/div[4]/div/div/div[3]/div[1]/div[2]/div/div/div/ul/li[1]/button'))).click()
        if brutto is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200202)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200302)"]', podatek)
        if doplata is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200402)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200502)"]', koscielny)
        zapiszKlase6()
    except:
        pass

#-------------------------------------------------------------------------------------#
# 045 KIND 
kindURL= 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageKind'
driver.get(kindURL)
try:
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="uebernehmenButton_451563490"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Cancel"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="closeButton_datenuebernahmeModal"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="anlagenAuswahlVertical"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="VAnlageKind"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="VAnlageSA"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Continue"]'))).click()
except:
    pass

time.sleep(.1)
ws = wb['ZP Status DE']
row = None
for i in range(1, ws.max_row + 1):
    full_name_cell = ws[f'A{i}'].value
    if full_name_cell is not None:
        full_name_parts = full_name_cell.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
            row = i
            break
def process_sentence(sentence):
    lines = sentence.split('\n')
    people = []
    for line in lines:
        # Podziel linię według spacji
        parts = line.split()
        # Przypisz odpowiednie fragmenty do zmiennych
        name = parts[0]
        date = parts[1]
        status = parts[2] + " " + parts[3]
        # Dodaj dane osoby do listy
        people.append((name, date, status))
    return people, len(people)
def remove_commas(text):
    return text.replace(",", "")
time.sleep(1)
kindURL= 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageKind'
driver.get(kindURL)
kind = ws[f'o{row}'].value
if kind is not None:
    kind = remove_commas(kind)
    people_data, number_of_people = process_sentence(kind)
    
    try:
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/MAVSAnlageKind[0]/VAnlageKind[*]"]'))).click()
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))).click()
    except:
        pass
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="JumpToPage/Startseite[0]/MAVSAnlageKind[0]/VAnlageKind[0]"]'))).click()
    #wypełnianie danych
    if number_of_people == 1:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="AngabenZumKind"]'))).click()
        for person in people_data:
            if person[2] == 'BEZ KG':
                pass
            else:
                wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500702)"]', person[2])
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500107)"]', person[0])
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500701)"]', person[1])
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_WohnortAusland(0)_fields(eruKindAng_KindWSAuslE0500704)"]', '01.01-31.12')
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500706)"]', 'Sachsen')

        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_WohnortAusland(0)_fields(eruKindAng_KindWSAuslE0500104)"]', 'POLEN')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageKind[0]/VAnlageKind[0]/AngabenZumKind[0]/WohnortAusland[0]"]'))).click()
        pyautogui.press('enter')
        str2KindURL= 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageKind/VAnlageKind/0/0633dfe3-9fb4-4baa-b163-223d550bfd06'
        driver.get(str2KindURL)
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_AE0500807)"]', 'leibliches Kind / Adoptivkind')
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_AE0500601)"]', '01.01-31.12')

        ws = wb['ZP Dane kont']
        for i in range(1, ws.max_row + 1):
            if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
                row = i
                break
        czyZonaty = ws[f'F{row}'].value
        if czyZonaty == "Żonaty":
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_BE0500808)"]', 'leibliches Kind / Adoptivkind')
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_BE0500805)"]', '01.01-31.12')
        pyautogui.press('enter')
#-------------------------------------------------------------------------------------#
# 055 Anlage N
# Wprowadzenie danych
# Przejście do sekcji Anlage N

AnlageN = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/AngabenZumArbeitslohn'
driver.get(AnlageN)

ws = wb['ZP Status DE']

row = None
for i in range(1, ws.max_row + 1):
    full_name_cell = ws[f'A{i}'].value
    if full_name_cell is not None:
        full_name_parts = full_name_cell.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
            row = i
            break

klasa = ws[f'AA{row}'].value
kl1 = ws[f'AA{row}'].value
brutto = ws[f'AB{row}'].value
podatek = ws[f'AC{row}'].value
doplata = ws[f'AD{row}'].value
koscielny = ws[f'AE{row}'].value
kurzarbeitgeld = ws[f'AF{row}'].value

#Usunięcie wszystkich danych klasa 6
try:
    element = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[*]"]'))
    )
    element.click()
except TimeoutException:
    pass
try:
    element = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))
    )
    element.click()
except TimeoutException:
    pass
#Usunięcie wszystkich danych klasa 6
try:
    element = WebDriverWait(driver, 1).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[*]"]'))
    )
    element.click()
except TimeoutException:
    pass
try:
    element = WebDriverWait(driver, 1).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))
    )
    element.click()
except TimeoutException:
    pass
kl1 = int(kl1)
if kl1 < 6:
    klasa1_pr1()
else:
    klasa6_pr1()

# Pobieranie i wprowadzanie dodatkowych danych
pit2 = ws[f'AG{row}'].value
if pit2 is not None:
    klasa = ws[f'AG{row}'].value
    kl2 = ws[f'AG{row}'].value
    brutto = ws[f'AH{row}'].value
    podatek = ws[f'AI{row}'].value
    doplata = ws[f'AJ{row}'].value
    koscielny = ws[f'AK{row}'].value
    kurzarbeitgeld = ws[f'AL{row}'].value
    kl2 = int(kl2)
    if kl2 < 6:
        if kl1 != kl2:
            klasa1_pr1()
        else:
            klasa1_pr2()
    else:
        if kl2 != kl1:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[0]"]'))).click()
            klasa6_pr1()
        else:
            klasa6_pr2()
pit3 = ws[f'AM{row}'].value
if pit3 is not None:
    klasa = ws[f'AM{row}'].value
    kl3 = ws[f'AM{row}'].value
    brutto = ws[f'AN{row}'].value
    podatek = ws[f'AO{row}'].value
    doplata = ws[f'AP{row}'].value
    koscielny = ws[f'AQ{row}'].value
    kurzarbeitgeld = ws[f'AR{row}'].value
    kl3 =int(kl3)
    if klasa < 6:
        if kl3==kl2 and kl3==kl1:
            klasa1_pr3()
        elif kl3!=kl2 and kl3==kl1:
            klasa1_pr2()
        elif kl3==kl2 and kl3!=kl1:
            klasa1_pr2()
        else:
            klasa1_pr1()
    else:
        if kl2==kl3:
            klasa6_pr2()
        elif kl1==kl3:
            klasa6_pr2()
        else:
            klasa6_pr1()

fahrkosten = ws[f'I{row}'].value
if fahrkosten is not None:
    fahrkostenUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(fahrkostenUrl)
    fahrkostenUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(fahrkostenUrl)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtFahrtkosten(0)_fields(eruNWkAWTFahrtE0205003)"]', 'Fahrkosten')
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtFahrtkosten(0)_fields(eruNWkAWTFahrtE0205004)"]', fahrkosten)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/ReisekostenAuswaertstaetigkeit[0]/AwtFahrtkosten[0]"]'))).click()

ubernachtungskosten = ws[f'J{row}'].value
if ubernachtungskosten is not None:
    ubernachtungskostenUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(ubernachtungskostenUrl)
    time.sleep(2)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtUebernachtungskosten(0)_fields(eruNWkAWTUebernachtE0206301)"]', 'Ubernachtungskosten')
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtUebernachtungskosten(0)_fields(eruNWkAWTUebernachtE0206302)"]', ubernachtungskosten)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/ReisekostenAuswaertstaetigkeit[0]/AwtUebernachtungskosten[0]"]'))).click()

wKabinie = ws[f'M{row}'].value
if wKabinie is not None:
    wKabinieURL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(wKabinieURL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_fields(eruNWkAWTPB_KraftfE0206501)"]', wKabinie)
    pyautogui.press('enter')

h24 = ws[f'K{row}'].value
if h24 is not None:
    h24URL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(h24URL)
    time.sleep(2)
    driver.get(h24URL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAInlE0205409)"]', h24)
    pyautogui.press('enter')

h8 = ws[f'L{row}'].value
if h8 is not None:
    h8URL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(h8URL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAInlE0205201)"]', h8)
    pyautogui.press('enter')

abUndAb = ws[f'N{row}'].value
if abUndAb is not None:
    abUndAbURL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(abUndAbURL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAInlE0205302)"]', abUndAb)
    pyautogui.press('enter')

pracodawca = ws[f'AT{row}'].value
if pracodawca is not None:
    pracodawcaURL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(pracodawcaURL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAVMA_ErsatzE0205108)"]', pracodawca)
    pyautogui.press('enter')


#Kirchensteuer
kosc1 = ws[f'AE{row}'].value
kosc2 = ws[f'AK{row}'].value
kosc3 = ws[f'AQ{row}'].value
if kosc1 is None:
    kosc1=0
if kosc2 is None:
    kosc2=0
if kosc3 is None:
    kosc3=0
kosc1 = int(kosc1)
kosc2 = int(kosc2)
kosc3 = int(kosc3)
Kirchensteuer = kosc1+kosc2+kosc3
if Kirchensteuer != 0:
    KirchensteuerUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageSA/Kirchensteuer'
    driver.get(KirchensteuerUrl)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/VAnlageSA[0]/Kirchensteuer[0]/KiSt[*]"]'))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))).click()
    except:
        pass
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageSA(0)_Kirchensteuer(0)_KiSt(0)_fields(eruSAKiStGezahltEinzE0108004)"]', Kirchensteuer)
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageSA(0)_Kirchensteuer(0)_KiSt(0)_fields(eruSAKiStGezahltEinzE0108003)"]', 'Kirchensteuer laut Lohnsteuerbescheinigung steuerpflichtige Person / Ehemann / Person A')
    pyautogui.press('enter')

nr22 = ws[f'W{row}'].value
nr23 = ws[f'X{row}'].value
if nr22 is not None or nr23 is not None:
    nr22i23URL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageVor/BeitraegeZurAltersvorsorge'
    driver.get(nr22i23URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="delete_btn_readMode_Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbLoeschenModal"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edit_btn_Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)"]'))).click()
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)_fields(eruVORAVorE2000401)"]', nr23)
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)_fields(eruVORAVorE2000801)"]', nr22)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="UpdateMzbItem/Startseite[0]/VAnlageVor[0]/BeitraegeZurAltersvorsorge[0]/BeitraegeZurAltersvorsorgeMZB[0]"]'))).click()

nr25 = ws[f'y{row}'].value
nr26 = ws[f'z{row}'].value
if nr25 is not None:
    nr25i26URL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageVor/BeitraegeInlGesKrankenPflegevers'
    driver.get(nr25i26URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="delete_btn_Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbLoeschenModal"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edit_btn_Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)"]'))).click()
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)_fields(eruVORBeitr_g_KV_PV_InlANE2001203)"]', nr25)
    if nr26 is None:
        nr26=0
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)_fields(eruVORBeitr_g_KV_PV_InlANE2001505)"]', nr26)
    pyautogui.press('enter')

nr27 = ws[f'AS{row}'].value
if nr27 is not None:
    nr27URL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageVor/WeitereSonstigeVorsorgeaufwendungen'
    driver.get(nr27URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[5]/main/div/div[4]/div/div/div[1]/div/div[2]/div[2]/div/div/div[3]/div/input[2]'))).click()
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.press('backspace')
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_WeitereSonstigeVorsorgeaufwendungen(0)_WeitSonsVorAW(0)_fields(eruVORWeit_Sons_VorAWPersE2004403)"]', nr27)
    pyautogui.press('enter')

#KRANKENGELD
krankengeld = ws[f'AU{row}'].value
if krankengeld is not None:
    krankengeldURL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VHauptvordruck/Einkommensersatzleistungen'
    driver.get(krankengeldURL)
    try:
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="delete_btn_Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)"]'))).click()
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbLoeschenModal"]'))).click()
    except:
        pass
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edit_btn_Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)"]'))).click()
    wait_and_send_keys('//*[@id="Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)_EinkErsInl(0)_fields(eruESt1AEink_ErsInlEinzE0104110)"]', 'Krankengeld')
    wait_and_send_keys('//*[@id="Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)_EinkErsInl(0)_fields(eruESt1AEink_ErsInlEinzE0104113)"]', krankengeld)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/VHauptvordruck[0]/Einkommensersatzleistungen[0]/MZBEinkErs[0]/EinkErsInl[0]"]'))).click()


#ZAROBKI W POLSCE WA-EST

WaEstUrl = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageWAESt'
time.sleep(1)
driver.get(WaEstUrl)


ZarobkiMezaNiem = ws[f'U{row}'].value
ZarobkiMezaNiem = int(ZarobkiMezaNiem)
ZarobkiZonyNiem = ws[f'V{row}'].value
checkbox1 = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]')
checkbox2 = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]')
    
if ZarobkiZonyNiem is not None:
    ZarobkiZonyNiem = int(ZarobkiZonyNiem)
    if ZarobkiMezaNiem + ZarobkiZonyNiem < 20000:
        if checkbox1.is_selected():
            if checkbox2.is_selected():
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]'))).click()
        else:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]'))).click()
    else:
        if checkbox2.is_selected():
            if checkbox1.is_selected():
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]'))).click()
        else:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]'))).click()
else:
    if ZarobkiMezaNiem < 20000:
        if checkbox1.is_selected():
            if checkbox2.is_selected():
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]'))).click()
        else:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]'))).click()
    else:
        if checkbox2.is_selected():
            if checkbox1.is_selected():
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]'))).click()
        else:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]'))).click()

try:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105303)"]'))).clear()
except:
    pass
time.sleep(1)
input_field = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105301)"]')
input_field.clear()
time.sleep(.5)
wait_and_send_keys('//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105301)"]', ZarobkiMezaNiem)
if ZarobkiZonyNiem is not None:
    ZarobkiZonyNiem = int(ZarobkiZonyNiem)
    input_field = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105303)"]')
    input_field.clear()
    time.sleep(.5)
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105303)"]', ZarobkiZonyNiem)
time.sleep(1)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="SwitchModusPruefen"]'))).click()
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="steuerberechnungLink"]'))).click()
input("Wciśnij Enter, aby zakończyć działanie programu...")



