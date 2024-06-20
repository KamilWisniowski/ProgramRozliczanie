import time
from selenium import webdriver
import pyautogui
from openpyxl import load_workbook
import tkinter as tk
from tkinter import simpledialog
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
#Przejście zakładkami na kolejny formularz
def KolejnyFormularz():
    time.sleep(.1)
    window_handles = driver.window_handles
    # Przełączenie się na pierwsze otwarte okno
    driver.switch_to.window(window_handles[1])
    time.sleep(.2)
    driver.close()
    time.sleep(.2)
    driver.switch_to.window(window_handles[0])
    time.sleep(.2)
    newurl='https://www.formulare-bfinv.de/ffw/content.do'
    driver.get(newurl)

def wait_and_send_keys(xpath, value):
    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    element.send_keys(value)

#Ścieżka do wyboru dokumentu
def PierwszaZgoda():
    driver.maximize_window()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="datenschutz"]/button'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/div[2]/ul/li[3]/div/div/div[1]/a/div/div[2]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[5]/div/a/span'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[6]/div/a/span'))).click()
# Funkcja do pobierania danych z okienka dialogowego
def get_user_input(prompt):
    root = tk.Tk()
    root.withdraw()  # Ukryj główne okno
    user_input = simpledialog.askstring(title="Dane użytkownika", prompt=prompt)
    return user_input

#POBIERANIE 
def PobieranieFormularza():
    time.sleep(.3)
    Printout = driver.find_element("xpath",'/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[3]/input').click()
    time.sleep(3)

    ShowPrintout = driver.find_element("xpath",'/html/body/div[4]/form/div[7]/div/div[1]/div/div[2]/p/a').click()
    time.sleep(1)

    pyautogui.hotkey('ctrl', 's')
    time.sleep(2)

    pyautogui.click(x=653, y=73)
    time.sleep(1)

    save_path = "C:\\Users\\Kamil\\Downloads"
    pyautogui.typewrite(save_path)
    time.sleep(.5)
    pyautogui.press('enter')
    time.sleep(.5)
    pyautogui.click(x=992, y=633)
    time.sleep(1)

# Pobieranie imienia i nazwiska
imie = get_user_input("Podaj imię:")
nazwisko = get_user_input("Podaj nazwisko:")
imie = imie.upper().strip()
nazwisko = nazwisko.upper().strip()

#WYPEŁNIANIE FORMULARZA
wb = load_workbook('Dane1.xlsx')
ws = wb['ZP Dane kont']

#Włączenie strony do rozliczania
driver = webdriver.Chrome()
url='https://www.formulare-bfinv.de/ffw/form/display.do?%24context=2802C5863D1DB0B5962F'
driver.get(url)
row = None
for i in range(1, ws.max_row + 1):
    if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
        row = i
        break
if row is None:
    print("Nie znaleziono danych dla podanej osoby.")
    driver.quit()
    exit()
stnum = ws[f'J{row}'].value
def WypelnijImieNazwiskoST():
    name = driver.find_element("xpath", '//*[@id="name"]')
    name.send_keys(nazwisko)
    time.sleep(.1)
    vorname = driver.find_element("xpath", '//*[@id="vorname"]')
    vorname.send_keys(imie)
    time.sleep(.1)
    if stnum is not None:
            driver.find_element("xpath",'//*[@id="steuernummer"]').send_keys(stnum)
    
def parse_address(address):
    pattern = re.compile(r'(?P<street>[\w\sąćęłńóśźżĄĆĘŁŃÓŚŹŻ]+)\s+(?P<block>\d+)(?:\s*(?P<part>[A-Z]))?(?:/(?P<apartment>\d+))?')
    match = pattern.match(address.strip())
    if match:
        return match.groupdict()
    else:
        raise ValueError("Format adresu jest niepoprawny")

#-------------------------------------------------------------------------------------#
#EST 1A
if 1==1:
    #Włączenie strony EST1a
    PierwszaZgoda()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[2]/div[1]/a/span[1]/span'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[4]/div/div/div/div[4]/input'))).click()
    time.sleep(2)

    #Wypełnianie danych EST1a
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="k1"]'))).click()
    WypelnijImieNazwiskoST()
    finanzamt = ws[f'I{row}'].value
    wait_and_send_keys('//*[@id="finanzamt"]', finanzamt)

    # Podział numeru identyfikacyjnego na części
    numer = str(ws[f'K{row}'].value)
    numer.strip()
    part1 = numer[:2]
    part2 = numer[2:5]
    part3 = numer[5:8]
    part4 = numer[8:]

    # Wypełnianie pól identyfikacyjnych
    wait_and_send_keys('//*[@id="identifikationssnummer"]', part1)
    wait_and_send_keys('//*[@id="identifikationssnummer2"]', part2)
    wait_and_send_keys('//*[@id="identifikationssnummer3"]', part3)
    wait_and_send_keys('//*[@id="identifikationssnummer4"]', part4)
    time.sleep(2)
    # Data urodzenia
    data_urodzenia = ws[f'N{row}'].value
    if data_urodzenia is not None:
        data_urodzenia_str = data_urodzenia.strftime('%d%m%Y')
        wait_and_send_keys('//*[@id="geburtsdatum"]', data_urodzenia_str)
    pyautogui.press('enter')
    pyautogui.press('enter')
    time.sleep(6)
    #Religia
    Religia = ws[f'O{row}'].value
    wait_and_send_keys('//*[@id="religion_barrierearm"]', Religia)
    pyautogui.press('enter')

    # Ulica i numer
    adres = ws[f'P{row}'].value
    if adres is not None:
        adres = adres.strip()
        adres = adres.upper()
        address_components = parse_address(adres)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="strasse_hausnummer"]'))).send_keys(address_components['street'])
        if address_components['block'] is not None:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="hausnummer"]'))).send_keys(address_components['block'])
        if address_components['apartment'] is not None:    
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="hausnummerzusatz"]'))).send_keys(address_components['apartment'])
        if address_components['part'] is not None:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ergaenzung"]'))).send_keys(address_components['part'])
    else:
        print("Adres jest pusty dla podanej osoby.")

    #data Ślubu i #Czy rozliczani razem
    DataSlubu = ws[f'R{row}'].value
    if DataSlubu is not None:
        DataSlubuS = DataSlubu.strftime('%d%m%Y')
        wait_and_send_keys('//*[@id="datum"]', DataSlubuS)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="k9"]'))).click()
    pyautogui.press('enter')
    time.sleep(4)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="staat2"]'))).send_keys("POLEN")

    Miejscowosc=ws[f'Q{row}'].value
    kod_pocztowy, miejscowosc = Miejscowosc.split(' ', 1)
    Miejscowosc = Miejscowosc.upper()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="wohnort"]'))).send_keys(miejscowosc)
    kod_pocztowy = kod_pocztowy.upper()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="postleitzahl4"]'))).send_keys(kod_pocztowy)

    #Dane żony
    if DataSlubu is not None:
        zName = driver.find_element("xpath", '//*[@id="abweichender_name"]')
        zName.send_keys(nazwisko)
        zVorname=ws[f'T{row}'].value
        zVorname = zVorname.upper()
        driver.find_element("xpath", '//*[@id="vorname_ehefrau"]').send_keys(zVorname)
        

        idNrZony = ws[f'L{row}'].value
        time.sleep(.2)
        if idNrZony is not None:
            # Podział numeru identyfikacyjnego na części
            znumer = str(ws[f'L{row}'].value)
            zpart1 = znumer[:2]
            zpart2 = znumer[2:5]
            zpart3 = znumer[5:8]
            zpart4 = znumer[8:]

            # Wypełnianie pól identyfikacyjnych
            zidentifikationssnummer5 = driver.find_element("xpath", '//*[@id="identifikationssnummer5"]')
            zidentifikationssnummer5.send_keys(zpart1)

            zidentifikationssnummer6 = driver.find_element("xpath", '//*[@id="identifikationssnummer6"]')
            zidentifikationssnummer6.send_keys(zpart2)

            zidentifikationssnummer7 = driver.find_element("xpath", '//*[@id="identifikationssnummer7"]')
            zidentifikationssnummer7.send_keys(zpart3)

            zidentifikationssnummer8 = driver.find_element("xpath", '//*[@id="identifikationssnummer8"]')
            zidentifikationssnummer8.send_keys(zpart4)
        #Religia
        zReligia=driver.find_element("xpath", '//*[@id="religion2_barrierearm"]')
        zReligia.send_keys(ws[f'O{row}'].value)
        # Data urodzenia
        zdata_urodzenia = ws[f'S{row}'].value
        if zdata_urodzenia is not None:
            zdata_urodzenia_str = zdata_urodzenia.strftime('%d%m%Y')
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="geburtsdatum2"]'))).send_keys(zdata_urodzenia_str)
        pyautogui.press('enter')
        time.sleep(4)

    #2 STRONA
    Strona2 = driver.find_element("xpath","/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[10]").click()
    time.sleep(3)

    # Podział numeru konta bankowego
    numer = str(ws[f'G{row}'].value)
    part1 = numer[:5]
    part2 = numer[5:10]
    part3 = numer[10:15]
    part4 = numer[15:20]
    part5 = numer[20:25]
    part6 = numer[25:30]
    part7 = numer[30:]

    # Wypełnianie pól identyfikacyjnych
    iban = driver.find_element("xpath", '//*[@id="iban2"]')
    iban.send_keys(part1)

    iban2 = driver.find_element("xpath", '//*[@id="iban8"]')
    iban2.send_keys(part2)

    iban3 = driver.find_element("xpath", '//*[@id="iban9"]')
    iban3.send_keys(part3)

    iban4 = driver.find_element("xpath", '//*[@id="iban10"]')
    iban4.send_keys(part4)

    iban5 = driver.find_element("xpath", '//*[@id="iban11"]')
    iban5.send_keys(part5)

    iban6 = driver.find_element("xpath", '//*[@id="iban13"]')
    iban6.send_keys(part6)

    iban7 = driver.find_element("xpath", '//*[@id="iban14"]')
    iban7.send_keys(part7)

    swift = ws[f'H{row}'].value
    if swift is not None:
        swift = driver.find_element("xpath", '//*[@id="bic"]')
        swift.send_keys(ws[f'H{row}'].value)
    time.sleep(.3)
    driver.find_element("xpath", '//*[@id="k10"]').click()
    time.sleep(.3)

    #Zmiana formularza na ZP Status De
    ws = wb['ZP Status DE']
    row = None
    for i in range(1, ws.max_row + 1):
        # Odczytanie danych z komórki A
        full_name_cell = ws[f'A{i}'].value
        # Sprawdzenie, czy komórka nie jest pusta
        if full_name_cell is not None:
            # Rozdzielenie nazwiska i imienia na podstawie spacji
            full_name_parts = full_name_cell.split()
            # Sprawdzenie, czy znaleźliśmy dopasowanie nazwiska i imienia
            if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
                row = i
                break
            
    #KRANKENGELD
    krankengeld = ws[f'AU{row}'].value
    if krankengeld is not None:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="k_hinweis"]'))).click()
        time.sleep(1)
        wait_and_send_keys('//*[@id="betrag8"]', krankengeld)

    #Pobieranie
    PobieranieFormularza()

sonderausgaben1 = ws[f'AE{row}'].value
sonderausgaben2 = ws[f'Ak{row}'].value
sonderausgaben3 = ws[f'Aq{row}'].value
if sonderausgaben1 is not None or sonderausgaben2 is not None or sonderausgaben3 is not None: 
    if sonderausgaben1 is not None:
        sonderausgaben1 = round(int(sonderausgaben1))
    if sonderausgaben2 is not None:
        sonderausgaben2 = round(int(sonderausgaben2))
    if sonderausgaben3 is not None:
        sonderausgaben3 = round(int(sonderausgaben3))
    KolejnyFormularz()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[5]/div[1]/a'))).click()
    WypelnijImieNazwiskoST()
    sonderausgabenCaly = sonderausgaben1+sonderausgaben2+sonderausgaben3
    wait_and_send_keys('//*[@id="eur"]', sonderausgabenCaly)
    PobieranieFormularza()

#-------------------------------------------------------------------------------------#
# 017 WA-EST
if 1==1:
    #Otwieranie strony WA-est
    KolejnyFormularz()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[7]/div[5]/div[2]/div[2]/div[2]/ul/li[9]/div[1]/a'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[7]/div/div[1]/div/div[4]/input'))).click()
    time.sleep(3)
    #Wypełnianie danych WA-est
    WypelnijImieNazwiskoST()

    #Zmiana formularza na ZP Status De
    ws = wb['ZP Status DE']
    row = None
    for i in range(1, ws.max_row + 1):
        # Odczytanie danych z komórki A
        full_name_cell = ws[f'A{i}'].value
        # Sprawdzenie, czy komórka nie jest pusta
        if full_name_cell is not None:
            # Rozdzielenie nazwiska i imienia na podstawie spacji
            full_name_parts = full_name_cell.split()
            # Sprawdzenie, czy znaleźliśmy dopasowanie nazwiska i imienia
            if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
                row = i
                break

    ZarobkiMezaNiem = ws[f'U{row}'].value
    ZarobkiZonyNiem = ws[f'V{row}'].value
    time.sleep(1)
    driver.find_element("xpath", '//*[@id="betrag3"]').send_keys(ZarobkiMezaNiem)
    time.sleep(1)
    if ZarobkiZonyNiem is not None:
        driver.find_element("xpath", '//*[@id="betrag4"]').send_keys(ZarobkiZonyNiem)
        time.sleep(1)

    ZarobkiMezaNiem = int(ZarobkiMezaNiem)
    ZarobkiZonyNiem = int(ZarobkiZonyNiem)
    if ZarobkiMezaNiem + ZarobkiZonyNiem < 20000:
        driver.find_element("xpath", '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[35]/input').click()
    else:
        driver.find_element("xpath", '/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[119]/input').click()
    PobieranieFormularza()
#-------------------------------------------------------------------------------------#
def process_sentence(sentence):
    lines = sentence.split('\n')
    people = []
    for line in lines:
        # Podziel linię według spacji
        parts = line.split()
        # Przypisz odpowiednie fragmenty do zmiennych
        name = parts[0]
        date = parts[1]
        status = parts[2] + " " + parts[3]
        # Dodaj dane osoby do listy
        people.append((name, date, status))
    return people, len(people)
def remove_commas(text):
    return text.replace(",", "")
# 045 KIND 
kind = ws[f'O{row}'].value
kind = remove_commas(kind)
if kind is not None:
    people_data, number_of_people = process_sentence(kind)
    KolejnyFormularz()
    time.sleep(1)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[17]/div[1]/a'))).click()
    time.sleep(2)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[4]/div/div/div/div[4]/input'))).click()
    time.sleep(3)
    if number_of_people == 1:
        WypelnijImieNazwiskoST()
        time.sleep(1)
        wait_and_send_keys('//*[@id="lfd_nr"]', '1')
        for person in people_data:
            if person[2] == 'BEZ KG':
                pass
            else:
                wait_and_send_keys('//*[@id="betrag"]', person[2])
        wait_and_send_keys('//*[@id="vorname2"]', person[0])
        wait_and_send_keys('//*[@id="datum"]', person[1])
        pyautogui.press('enter')
        time.sleep(6)
        wait_and_send_keys('//*[@id="datum5"]', '0101')
        pyautogui.press('enter')
        time.sleep(5)
        wait_and_send_keys('//*[@id="datum6"]', '3112')
        pyautogui.press('enter')
        time.sleep(5)
        wait_and_send_keys('//*[@id="staat"]', 'POLEN')
        ws = wb['ZP Dane kont']
        for i in range(1, ws.max_row + 1):
            if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
                row = i
                break
        wait_and_send_keys('//*[@id="k1"]', '1')
        czyZonaty = ws[f'F{row}'].value
        if czyZonaty == "Żonaty":
            wait_and_send_keys('//*[@id="k2"]', '1')
    PobieranieFormularza()



#-------------------------------------------------------------------------------------#
# 055 Anlage N
KolejnyFormularz()
if 1==1:
    #Otwieranie strony Anlage N
    AnlageN= driver.find_element("xpath", '/html/body/div[7]/div[5]/div[2]/div[2]/div[2]/ul/li[21]/div[1]/a').click()
    time.sleep(3)

    #Zgody
    Zgoda = driver.find_element("xpath","/html/body/div[4]/form/div[4]/div/div/div/div[4]/input").click()
    time.sleep(2)
    SteuerpflichtigePersonA= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[1]/input").click()
    WichtigerHinweis= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[4]/input").click()
    
    time.sleep(3)
    #Wypełnianie danych Osobowych
    name = driver.find_element("xpath", '//*[@id="name"]')
    name.send_keys(nazwisko)

    vorname = driver.find_element("xpath", '//*[@id="name2"]')
    vorname.send_keys(imie)

    ws = wb['ZP Dane kont']
    row = None
    for i in range(1, ws.max_row + 1):
        if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
            row = i
            break
    if row is None:
        print("Nie znaleziono danych dla podanej osoby.")
        driver.quit()
        exit()
    stnum = ws[f'J{row}'].value
    if stnum is not None:
        driver.find_element("xpath",'//*[@id="steuernummer"]').send_keys(stnum)
    time.sleep(1)

    ws = wb['ZP Status DE']
    # Szukanie osoby o podanym nazwisku i imieniu w arkuszu "ZP Status De"
    row = None
    for i in range(1, ws.max_row + 1):
        # Odczytanie danych z komórki A
        full_name_cell = ws[f'A{i}'].value
        # Sprawdzenie, czy komórka nie jest pusta
        if full_name_cell is not None:
            # Rozdzielenie nazwiska i imienia na podstawie spacji
            full_name_parts = full_name_cell.split()
            # Sprawdzenie, czy znaleźliśmy dopasowanie nazwiska i imienia
            if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
                row = i
                break

    pit1 = ws[f'AA{row}'].value
    if pit1 is not None:
        ws= wb['ZP Status DE']

        #Wypełnianie danych aNLAGE n
        #--Tutaj będą zarobki w niemczech
        klasa = ws[f'AA{row}'].value
        brutto1 = ws[f'AB{row}'].value
        if brutto1 is not None:
            brutto1 = round(int(ws[f'AB{row}'].value))
        podatek1 = ws[f'AC{row}'].value
        if podatek1 is not None:
            podatek1 = round(int(ws[f'AC{row}'].value))
        doplata1 = ws[f'AD{row}'].value
        if doplata1 is not None:                   
            doplata1 = round(int(ws[f'AD{row}'].value))
        koscielny1 = ws[f'AE{row}'].value
        if koscielny1 is not None:
            koscielny1 = round(int(ws[f'AE{row}'].value))
        kurzarbeitgeld1 = ws[f'AF{row}'].value
        if kurzarbeitgeld1 is not None:
            kurzarbeitgeld1 = round(int(ws[f'AF{row}'].value))

        if klasa < 6:
            driver.find_element("xpath",'//*[@id="steuerklasse"]').send_keys(klasa)
            pyautogui.press('enter')
            if brutto1 is not None:
                driver.find_element("xpath",'//*[@id="betrag"]').send_keys(brutto1)
            if podatek1 is not None:
                driver.find_element("xpath",'//*[@id="betrag2"]').send_keys(podatek1)
            if doplata1 is not None:
                driver.find_element("xpath",'//*[@id="betrag3"]').send_keys(doplata1)
            if koscielny1 is not None:
                driver.find_element("xpath",'//*[@id="betrag4"]').send_keys(koscielny1)
        else:
            if brutto1 is not None:    
                driver.find_element("xpath",'//*[@id="betrag6"]').send_keys(brutto1)
            if podatek1 is not None:
                driver.find_element("xpath",'//*[@id="betrag7"]').send_keys(podatek1)
            if doplata1 is not None:    
                driver.find_element("xpath",'//*[@id="betrag8"]').send_keys(doplata1)
            if koscielny1 is not None:
                driver.find_element("xpath",'//*[@id="betrag9"]').send_keys(koscielny1)

        if kurzarbeitgeld1 is not None:
            driver.find_element("xpath",'//*[@id="betrag30"]').send_keys(kurzarbeitgeld1)

        #Przejście na stronę 3
        Strona3= driver.find_element("xpath","/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[11]").click()
        time.sleep(4)

        BERUFSKLEIDUNg= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/input[19]")
        BERUFSKLEIDUNg.send_keys('BERUFSKLEIDUNG')
        BERUFSKLEIDUNg110= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/input[20]")
        BERUFSKLEIDUNg110.send_keys('110')
        Berechnen= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/input[22]").click()
        time.sleep(1)

        #Przejście na stronę 4
        Strona3 = driver.find_element("xpath","/html/body/div[4]/form/div[6]/div/div/div[1]/div[2]/div[1]/div/div[12]").click()
        time.sleep(3)

        fahrtkosten = ws[f'I{row}'].value
        ubernachtungskosten = ws[f'J{row}'].value
        h24 = ws[f'K{row}'].value
        h8 = ws[f'L{row}'].value
        wKabinie = ws[f'M{row}'].value
        anUndAbrei = ws[f'N{row}'].value
        pracodawca = ws[f'AT{row}'].value

        if fahrtkosten is not None:
            driver.find_element("xpath",'//*[@id="fahrkosten"]').send_keys('Fahrtkosten')
            driver.find_element("xpath",'//*[@id="betrag45"]').send_keys(fahrtkosten)

        if ubernachtungskosten is not None:
            driver.find_element("xpath",'//*[@id="uebernachtungskosten"]').send_keys('Ubernachtungskosten')
            driver.find_element("xpath",'//*[@id="betrag71"]').send_keys(ubernachtungskosten)

        if h24 is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage3"]').send_keys(h24)

        if h8 is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage"]').send_keys(h8)

        if wKabinie is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage4"]').send_keys(wKabinie)

        if anUndAbrei is not None:
            driver.find_element("xpath",'//*[@id="zahl_der_tage2"]').send_keys(anUndAbrei)
        if pracodawca is not None:
            driver.find_element("xpath",'//*[@id="betrag47"]').send_keys(pracodawca)
        time.sleep(1)

        #Pobieranie
        PobieranieFormularza()
    pit2 = ws[f'AG{row}'].value
    if pit2 is not None:

        KolejnyFormularz()

        #Otwieranie strony Anlage N
        AnlageN= driver.find_element("xpath", '/html/body/div[7]/div[5]/div[2]/div[2]/div[2]/ul/li[21]/div[1]/a').click()
        time.sleep(3)

        #Zgody
        Zgoda = driver.find_element("xpath","/html/body/div[4]/form/div[4]/div/div/div/div[4]/input").click()
        time.sleep(2)
        SteuerpflichtigePersonA= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[1]/input").click()
        WichtigerHinweis= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[4]/input").click()
        time.sleep(2)

        time.sleep(3)
        #Wypełnianie danych Osobowych
        name = driver.find_element("xpath", '//*[@id="name"]')
        name.send_keys(nazwisko)

        vorname = driver.find_element("xpath", '//*[@id="name2"]')
        vorname.send_keys(imie)

        ws = wb['ZP Dane kont']
        row = None
        for i in range(1, ws.max_row + 1):
            if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
                row = i
                break
        if row is None:
            print("Nie znaleziono danych dla podanej osoby.")
            driver.quit()
            exit()
        stnum = ws[f'J{row}'].value
        if stnum is not None:
            driver.find_element("xpath",'//*[@id="steuernummer"]').send_keys(stnum)
        time.sleep(1)

        # Szukanie osoby o podanym nazwisku i imieniu w arkuszu "ZP Status De"
        row = None
        for i in range(1, ws.max_row + 1):
            # Odczytanie danych z komórki A
            full_name_cell = ws[f'A{i}'].value
            # Sprawdzenie, czy komórka nie jest pusta
            if full_name_cell is not None:
                # Rozdzielenie nazwiska i imienia na podstawie spacji
                full_name_parts = full_name_cell.split()
                # Sprawdzenie, czy znaleźliśmy dopasowanie nazwiska i imienia
                if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
                    row = i
                    break

        #Wypełnianie danych aNLAGE n
        #--Tutaj będą zarobki w niemczech
        klasa = ws[f'AG{row}'].value
        brutto1 = ws[f'Ah{row}'].value
        if brutto1 is not None:
            brutto1 = round(int(ws[f'Ah{row}'].value))
        podatek1 = ws[f'Ai{row}'].value
        if podatek1 is not None:
            podatek1 = round(int(ws[f'Ai{row}'].value))
        doplata1 = ws[f'Aj{row}'].value
        if doplata1 is not None:                   
            doplata1 = round(int(ws[f'Aj{row}'].value))
        koscielny1 = ws[f'Ak{row}'].value
        if koscielny1 is not None:
            koscielny1 = round(int(ws[f'Ak{row}'].value))
        kurzarbeitgeld1 = round(int(ws[f'Al{row}'].value))

        if klasa < 6:
            driver.find_element("xpath",'//*[@id="steuerklasse"]').send_keys(klasa)
            pyautogui.press('enter')
            if brutto1 is not None:
                driver.find_element("xpath",'//*[@id="betrag"]').send_keys(brutto1)
            if podatek1 is not None:
                driver.find_element("xpath",'//*[@id="betrag2"]').send_keys(podatek1)
            if doplata1 is not None:
                driver.find_element("xpath",'//*[@id="betrag3"]').send_keys(doplata1)
            if koscielny1 is not None:
                driver.find_element("xpath",'//*[@id="betrag4"]').send_keys(koscielny1)
        else:
            if brutto1 is not None:    
                driver.find_element("xpath",'//*[@id="betrag6"]').send_keys(brutto1)
            if podatek1 is not None:
                driver.find_element("xpath",'//*[@id="betrag7"]').send_keys(podatek1)
            if doplata1 is not None:    
                driver.find_element("xpath",'//*[@id="betrag8"]').send_keys(doplata1)
            if koscielny1 is not None:
                driver.find_element("xpath",'//*[@id="betrag9"]').send_keys(koscielny1)

        if kurzarbeitgeld1 is not None:
            driver.find_element("xpath",'//*[@id="betrag30"]').send_keys(kurzarbeitgeld1)
        #Pobieranie
        PobieranieFormularza()
    pit3 = ws[f'AM{row}'].value
    if pit3 is not None:
        KolejnyFormularz()

        #Otwieranie strony Anlage N
        AnlageN= driver.find_element("xpath", '/html/body/div[7]/div[5]/div[2]/div[2]/div[2]/ul/li[21]/div[1]/a').click()
        time.sleep(3)

        #Zgody
        Zgoda = driver.find_element("xpath","/html/body/div[4]/form/div[4]/div/div/div/div[4]/input").click()
        time.sleep(2)
        SteuerpflichtigePersonA= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[1]/input").click()
        WichtigerHinweis= driver.find_element("xpath","/html/body/div[4]/form/div[7]/div/div[4]/div/div[1]/div/div/div/div/div[4]/input").click()
        time.sleep(2)

        time.sleep(3)
        #Wypełnianie danych Osobowych
        name = driver.find_element("xpath", '//*[@id="name"]')
        name.send_keys(nazwisko)

        vorname = driver.find_element("xpath", '//*[@id="name2"]')
        vorname.send_keys(imie)

        ws = wb['ZP Dane kont']
        row = None
        for i in range(1, ws.max_row + 1):
            if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
                row = i
                break
        if row is None:
            print("Nie znaleziono danych dla podanej osoby.")
            driver.quit()
            exit()
        stnum = ws[f'J{row}'].value
        if stnum is not None:
            driver.find_element("xpath",'//*[@id="steuernummer"]').send_keys(stnum)
        time.sleep(1)
        ws= wb['ZP Status DE']

        # Szukanie osoby o podanym nazwisku i imieniu w arkuszu "ZP Status De"
        row = None
        for i in range(1, ws.max_row + 1):
            # Odczytanie danych z komórki A
            full_name_cell = ws[f'A{i}'].value
            # Sprawdzenie, czy komórka nie jest pusta
            if full_name_cell is not None:
                # Rozdzielenie nazwiska i imienia na podstawie spacji
                full_name_parts = full_name_cell.split()
                # Sprawdzenie, czy znaleźliśmy dopasowanie nazwiska i imienia
                if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
                    row = i
                    break

        #Wypełnianie danych aNLAGE n
        #--Tutaj będą zarobki w niemczech
        klasa = ws[f'AM{row}'].value
        brutto1 = ws[f'AB{row}'].value
        if brutto1 is not None:
            brutto1 = round(int(ws[f'AB{row}'].value))
        podatek1 = ws[f'AC{row}'].value
        if podatek1 is not None:
            podatek1 = round(int(ws[f'AC{row}'].value))
        doplata1 = ws[f'AD{row}'].value
        if doplata1 is not None:                   
            doplata1 = round(int(ws[f'AD{row}'].value))
        koscielny1 = ws[f'AE{row}'].value
        if koscielny1 is not None:
            koscielny1 = round(int(ws[f'AE{row}'].value))
        kurzarbeitgeld1 = round(int(ws[f'AF{row}'].value))

        if klasa < 6:
            driver.find_element("xpath",'//*[@id="steuerklasse"]').send_keys(klasa)
            pyautogui.press('enter')
            if brutto1 is not None:
                driver.find_element("xpath",'//*[@id="betrag"]').send_keys(brutto1)
            if podatek1 is not None:
                driver.find_element("xpath",'//*[@id="betrag2"]').send_keys(podatek1)
            if doplata1 is not None:
                driver.find_element("xpath",'//*[@id="betrag3"]').send_keys(doplata1)
            if koscielny1 is not None:
                driver.find_element("xpath",'//*[@id="betrag4"]').send_keys(koscielny1)
        else:
            if brutto1 is not None:    
                driver.find_element("xpath",'//*[@id="betrag6"]').send_keys(brutto1)
            if podatek1 is not None:
                driver.find_element("xpath",'//*[@id="betrag7"]').send_keys(podatek1)
            if doplata1 is not None:    
                driver.find_element("xpath",'//*[@id="betrag8"]').send_keys(doplata1)
            if koscielny1 is not None:
                driver.find_element("xpath",'//*[@id="betrag9"]').send_keys(koscielny1)

        if kurzarbeitgeld1 is not None:
            driver.find_element("xpath",'//*[@id="betrag30"]').send_keys(kurzarbeitgeld1)
        #Pobieranie
        PobieranieFormularza()

#-------------------------------------------------------------------------------------#
#Vorsorgeaufwendungen
KolejnyFormularz()
#Otwieranie strony Vorsorgeaufwendungen
vorsorgeaufwendungen= driver.find_element("xpath", '//*[@id="lip_formcatalog"]/div[2]/div[2]/div[2]/ul/li[35]/div[1]/a').click()
time.sleep(3)

#Zgoda nr1
driver.find_element("xpath",'/html/body/div[4]/form/div[4]/div/div/div/div[4]/input').click()
time.sleep(.5)
ws = wb['ZP Dane kont']
row = None
for i in range(1, ws.max_row + 1):
    if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
        row = i
        break
if row is None:
    print("Nie znaleziono danych dla podanej osoby.")
    driver.quit()
    exit()
time.sleep(3)        
WypelnijImieNazwiskoST()
time.sleep(2)
#Zgoda nr2
driver.find_element("xpath",'//*[@id="k_hinweis"]').click()
time.sleep(3)

ws= wb['ZP Status DE']
# Szukanie osoby o podanym nazwisku i imieniu w arkuszu "ZP Status De"
row = None
for i in range(1, ws.max_row + 1):
    full_name_cell = ws[f'A{i}'].value
    if full_name_cell is not None:
        full_name_parts = full_name_cell.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
            row = i
            break

nr22 = ws[f'W{row}'].value
nr23 = ws[f'X{row}'].value
nr25 = ws[f'Y{row}'].value
nr26 = ws[f'Z{row}'].value
driver.find_element("xpath",'//*[@id="betrag"]').send_keys(nr23)
driver.find_element("xpath",'//*[@id="betrag5"]').send_keys(nr22)
driver.find_element("xpath",'//*[@id="betrag15"]').send_keys(nr25)
driver.find_element("xpath",'//*[@id="betrag18"]').send_keys(nr26)
nr27 = ws[f'AS{row}'].value
if nr27 is not None:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="lip_buttonPanel"]/div[1]/div/div[11]'))).click()
    wait_and_send_keys('//*[@id="betrag71"]', nr27)
time.sleep(.5)

#Pobieranie
PobieranieFormularza()

input("Wciśnij Enter, aby zakończyć działanie programu...")



