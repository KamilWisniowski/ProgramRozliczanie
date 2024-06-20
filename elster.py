import time
from selenium import webdriver
import pyautogui
from openpyxl import load_workbook
import tkinter as tk
from tkinter import simpledialog
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import re

# Funkcja do pobierania danych z okienka dialogowego
def get_user_input(prompt):
    root = tk.Tk()
    root.withdraw()  # Ukryj główne okno
    user_input = simpledialog.askstring(title="Dane użytkownika", prompt=prompt)
    return user_input

# Załadowanie EXCELA
wb = load_workbook('Dane1.xlsx')
ws = wb['ZP Status DE']

# Pobieranie imienia i nazwiska
imie = get_user_input("Podaj imię:")
nazwisko = get_user_input("Podaj nazwisko:")
imie = imie.upper().strip()
nazwisko = nazwisko.upper().strip()
row = None
for i in range(1, ws.max_row + 1):
    full_name_cell = ws[f'A{i}'].value
    if full_name_cell is not None:
        full_name_parts = full_name_cell.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
            row = i
            break

# Włączenie strony do rozliczania (Logowanie)
driver = webdriver.Chrome()
url = 'https://www.elster.de/eportal/login/softpse'
driver.get(url)
driver.maximize_window()
# Definicja funkcji do parsowania adresu
def parse_address(address):
    pattern = re.compile(r'(?P<street>[\w\sąćęłńóśźżĄĆĘŁŃÓŚŹŻ]+)\s+(?P<block>\d+)(?:/(?P<apartment>\d+)(?P<part>[AB]?))?')
    match = pattern.match(address.strip())
    if match:
        return match.groupdict()
    else:
        raise ValueError("Format adresu jest niepoprawny")

# Wybór metody logowania
driver.find_element(By.XPATH, '//*[@id="submitButton_loginBox.file_cert"]').click()
time.sleep(.5)
pyautogui.click(x=115, y=488)
pyautogui.click(x=412, y=195)
pyautogui.press('enter')

# Wprowadzenie hasła i zatwierdzenie
driver.find_element(By.XPATH, '//*[@id="password"]').send_keys('GrandaKamil11071989!')
driver.find_element(By.XPATH, '//*[@id="bestaetigenButton"]').click()

try:
    element = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="temporaereaufgaben_nein_button"]'))
    )
    element.click()


except TimeoutException:
    # Element nie pojawił się w przeciągu 10 sekund, więc przechodzimy dalej
    pass
# Oczekiwanie na załadowanie strony i przejście do formularzy

meineFormulareURL= 'https://www.elster.de/eportal/meineformulare'
driver.get(meineFormulareURL)
# Pobranie roku rozliczenia i przejście do odpowiedniego formularza
rokRoliczenia = ws[f'C{row}'].value
if rokRoliczenia == 2023:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="oeffneEntwurf_452354774"]'))).click()
elif rokRoliczenia == 2022:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="oeffneEntwurf_425366900"]'))).click()
elif rokRoliczenia == 2021:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="oeffneEntwurf_433741909"]'))).click()


def wait_and_send_keys(xpath, value):
    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    element.send_keys(value)

def zapiszKlase1():
    try:
        WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[0]"]'))).click()

    except:
        pass
    time.sleep(1)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[1]"]'))).click()
    except:
        pass
    time.sleep(1)
    try:
        WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[2]"]'))).click()
    except:
        pass

def zapiszKlase6():
    try:
        WebDriverWait(driver, .1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[0]"]'))).click()
    except:
        pass
    try:
        WebDriverWait(driver, .1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[1]"]'))).click()
    except:
        pass


def klasa1_pr1():
    try:
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_fields(eruNArbLLStB_1_5_SumE0200002)"]', klasa)
        pyautogui.press('enter')
        if brutto is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200204)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200304)"]', podatek)
        if doplata is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200404)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(0)_fields(eruNArbLLStB_1_5_EinzE0200504)"]', koscielny)
        zapiszKlase1()
    except:
        pass
def klasa1_pr2():
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="AddMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf"]'))).click()
    try:
        if brutto is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200204)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200304)"]', podatek)
        if doplata is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200404)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200504)"]', koscielny)
        zapiszKlase1()
    except:
        pass
def klasa1_pr3():
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[1]"]'))).click()
    WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="AddMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf"]'))).click()
    try:
        if brutto is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200204)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200304)"]', podatek)
        if doplata is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(2)_fields(eruNArbLLStB_1_5_EinzE0200404)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlassenEinsBisFuenf(1)_fields(eruNArbLLStB_1_5_EinzE0200504)"]', koscielny)
        zapiszKlase1()
    except:
        pass
def klasa6_pr1():
    try:
        if brutto is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200202)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200302)"]', podatek)
        if doplata is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200402)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(0)_fields(eruNArbLLStB_6_EinzE0200502)"]', koscielny)
        zapiszKlase6()
    except:
        pass
def klasa6_pr2():
    try:
        WebDriverWait(driver, 12).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[5]/main/div/div[4]/div/div/div[3]/div[1]/div[2]/div/div/div/ul/li[1]/button'))).click()
        if brutto is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200202)"]', brutto)
        if podatek is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200302)"]', podatek)
        if doplata is not None:    
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200402)"]', doplata)
        if koscielny is not None:
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_AngabenZumArbeitslohn(0)_LStBKlasseSechsUrlaubskasse(1)_fields(eruNArbLLStB_6_EinzE0200502)"]', koscielny)
        zapiszKlase6()
    except:
        pass

#-------------------------------------------------------------------------------------#
# 045 KIND 
kindURL= 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageKind'
driver.get(kindURL)
try:
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="uebernehmenButton_451563490"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Cancel"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="closeButton_datenuebernahmeModal"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="anlagenAuswahlVertical"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="VAnlageKind"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="VAnlageSA"]'))).click()
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Continue"]'))).click()
except:
    pass

time.sleep(.1)
ws = wb['ZP Status DE']
row = None
for i in range(1, ws.max_row + 1):
    full_name_cell = ws[f'A{i}'].value
    if full_name_cell is not None:
        full_name_parts = full_name_cell.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
            row = i
            break
def process_sentence(sentence):
    lines = sentence.split('\n')
    people = []
    for line in lines:
        # Podziel linię według spacji
        parts = line.split()
        # Przypisz odpowiednie fragmenty do zmiennych
        name = parts[0]
        date = parts[1]
        status = parts[2] + " " + parts[3]
        # Dodaj dane osoby do listy
        people.append((name, date, status))
    return people, len(people)
def remove_commas(text):
    return text.replace(",", "")
time.sleep(1)
kindURL= 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageKind'
driver.get(kindURL)
kind = ws[f'o{row}'].value
if kind is not None:
    kind = remove_commas(kind)
    people_data, number_of_people = process_sentence(kind)
    
    try:
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/MAVSAnlageKind[0]/VAnlageKind[*]"]'))).click()
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))).click()
    except:
        pass
    WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="JumpToPage/Startseite[0]/MAVSAnlageKind[0]/VAnlageKind[0]"]'))).click()
    #wypełnianie danych
    if number_of_people == 1:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="AngabenZumKind"]'))).click()
        for person in people_data:
            if person[2] == 'BEZ KG':
                pass
            else:
                wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500702)"]', person[2])
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500107)"]', person[0])
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500701)"]', person[1])
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_WohnortAusland(0)_fields(eruKindAng_KindWSAuslE0500704)"]', '01.01-31.12')
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_fields(eruKindAng_KindAllgE0500706)"]', 'Sachsen')

        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_AngabenZumKind(0)_WohnortAusland(0)_fields(eruKindAng_KindWSAuslE0500104)"]', 'POLEN')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageKind[0]/VAnlageKind[0]/AngabenZumKind[0]/WohnortAusland[0]"]'))).click()
        pyautogui.press('enter')
        str2KindURL= 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageKind/VAnlageKind/0/0633dfe3-9fb4-4baa-b163-223d550bfd06'
        driver.get(str2KindURL)
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_AE0500807)"]', 'leibliches Kind / Adoptivkind')
        wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_AE0500601)"]', '01.01-31.12')

        ws = wb['ZP Dane kont']
        for i in range(1, ws.max_row + 1):
            if ws[f'A{i}'].value == imie and ws[f'B{i}'].value == nazwisko:
                row = i
                break
        czyZonaty = ws[f'F{row}'].value
        if czyZonaty == "Żonaty":
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_BE0500808)"]', 'leibliches Kind / Adoptivkind')
            wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageKind(0)_VAnlageKind(0)_0633dfe3-9fb4-4baa-b163-223d550bfd06(0)_fields(eruKindK_VerhK_Verh_BE0500805)"]', '01.01-31.12')
        pyautogui.press('enter')
#-------------------------------------------------------------------------------------#
# 055 Anlage N
# Wprowadzenie danych
# Przejście do sekcji Anlage N

AnlageN = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/AngabenZumArbeitslohn'
driver.get(AnlageN)

ws = wb['ZP Status DE']

row = None
for i in range(1, ws.max_row + 1):
    full_name_cell = ws[f'A{i}'].value
    if full_name_cell is not None:
        full_name_parts = full_name_cell.split()
        if full_name_parts[0] == nazwisko and full_name_parts[1] == imie:
            row = i
            break

klasa = ws[f'AA{row}'].value
kl1 = ws[f'AA{row}'].value
brutto = ws[f'AB{row}'].value
podatek = ws[f'AC{row}'].value
doplata = ws[f'AD{row}'].value
koscielny = ws[f'AE{row}'].value
kurzarbeitgeld = ws[f'AF{row}'].value

#Usunięcie wszystkich danych klasa 6
try:
    element = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlassenEinsBisFuenf[*]"]'))
    )
    element.click()
except TimeoutException:
    pass
try:
    element = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))
    )
    element.click()
except TimeoutException:
    pass
#Usunięcie wszystkich danych klasa 6
try:
    element = WebDriverWait(driver, 1).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[*]"]'))
    )
    element.click()
except TimeoutException:
    pass
try:
    element = WebDriverWait(driver, 1).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))
    )
    element.click()
except TimeoutException:
    pass
if kl1 < 6:
    klasa1_pr1()
else:
    klasa6_pr1()

# Pobieranie i wprowadzanie dodatkowych danych
pit2 = ws[f'AG{row}'].value
if pit2 is not None:
    klasa = ws[f'AG{row}'].value
    kl2 = ws[f'AG{row}'].value
    brutto = ws[f'AH{row}'].value
    podatek = ws[f'AI{row}'].value
    doplata = ws[f'AJ{row}'].value
    koscielny = ws[f'AK{row}'].value
    kurzarbeitgeld = ws[f'AL{row}'].value
    if kl2 < 6:
        if kl1 != kl2:
            klasa1_pr1()
        else:
            klasa1_pr2()
    else:
        if kl2 != kl1:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/AngabenZumArbeitslohn[0]/LStBKlasseSechsUrlaubskasse[0]"]'))).click()
            klasa6_pr1()
        else:
            klasa6_pr2()
pit3 = ws[f'AM{row}'].value
if pit3 is not None:
    klasa = ws[f'AM{row}'].value
    kl3 = ws[f'AM{row}'].value
    brutto = ws[f'AN{row}'].value
    podatek = ws[f'AO{row}'].value
    doplata = ws[f'AP{row}'].value
    koscielny = ws[f'AQ{row}'].value
    kurzarbeitgeld = ws[f'AR{row}'].value
    if klasa < 6:
        if kl3==kl2 and kl3==kl1:
            klasa1_pr3()
        elif kl3!=kl2 and kl3==kl1:
            klasa1_pr2()
        elif kl3==kl2 and kl3!=kl1:
            klasa1_pr2()
        else:
            klasa1_pr1()
    else:
        if kl2==kl3:
            klasa6_pr2()
        elif kl1==kl3:
            klasa6_pr2()
        else:
            klasa6_pr1()

fahrkosten = ws[f'I{row}'].value
if fahrkosten is not None:
    fahrkostenUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(fahrkostenUrl)
    fahrkostenUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(fahrkostenUrl)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtFahrtkosten(0)_fields(eruNWkAWTFahrtE0205003)"]', 'Fahrkosten')
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtFahrtkosten(0)_fields(eruNWkAWTFahrtE0205004)"]', fahrkosten)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/ReisekostenAuswaertstaetigkeit[0]/AwtFahrtkosten[0]"]'))).click()

ubernachtungskosten = ws[f'J{row}'].value
if ubernachtungskosten is not None:
    ubernachtungskostenUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(ubernachtungskostenUrl)
    time.sleep(2)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtUebernachtungskosten(0)_fields(eruNWkAWTUebernachtE0206301)"]', 'Ubernachtungskosten')
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_AwtUebernachtungskosten(0)_fields(eruNWkAWTUebernachtE0206302)"]', ubernachtungskosten)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/MAVSAnlageN[0]/VAnlageN[0]/ReisekostenAuswaertstaetigkeit[0]/AwtUebernachtungskosten[0]"]'))).click()

wKabinie = ws[f'M{row}'].value
if wKabinie is not None:
    wKabinieURL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/ReisekostenAuswaertstaetigkeit'
    driver.get(wKabinieURL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_ReisekostenAuswaertstaetigkeit(0)_fields(eruNWkAWTPB_KraftfE0206501)"]', wKabinie)
    pyautogui.press('enter')

h24 = ws[f'K{row}'].value
if h24 is not None:
    h24URL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(h24URL)
    time.sleep(2)
    driver.get(h24URL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAInlE0205409)"]', h24)
    pyautogui.press('enter')

h8 = ws[f'L{row}'].value
if h8 is not None:
    h8URL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(h8URL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAInlE0205201)"]', h8)
    pyautogui.press('enter')

abUndAb = ws[f'N{row}'].value
if abUndAb is not None:
    abUndAbURL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(abUndAbURL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAInlE0205302)"]', abUndAb)
    pyautogui.press('enter')

pracodawca = ws[f'AT{row}'].value
if pracodawca is not None:
    pracodawcaURL='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/MAVSAnlageN/VAnlageN/0/PauschbetraegeVerpflegung'
    driver.get(pracodawcaURL)
    wait_and_send_keys('//*[@id="Startseite(0)_MAVSAnlageN(0)_VAnlageN(0)_PauschbetraegeVerpflegung(0)_fields(eruNWkVMAVMA_ErsatzE0205108)"]', pracodawca)
    pyautogui.press('enter')


#Kirchensteuer
kosc1 = ws[f'AE{row}'].value
kosc2 = ws[f'AK{row}'].value
kosc3 = ws[f'AQ{row}'].value
if kosc1 is None:
    kosc1=0
if kosc2 is None:
    kosc2=0
if kosc3 is None:
    kosc3=0
kosc1 = int(kosc1)
kosc2 = int(kosc2)
kosc3 = int(kosc3)
Kirchensteuer = kosc1+kosc2+kosc3
if Kirchensteuer != 0:
    KirchensteuerUrl='https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageSA/Kirchensteuer'
    driver.get(KirchensteuerUrl)
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ClearMzbItems/Startseite[0]/VAnlageSA[0]/Kirchensteuer[0]/KiSt[*]"]'))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbAlleLoeschenModal"]'))).click()
    except:
        pass
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageSA(0)_Kirchensteuer(0)_KiSt(0)_fields(eruSAKiStGezahltEinzE0108004)"]', Kirchensteuer)
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageSA(0)_Kirchensteuer(0)_KiSt(0)_fields(eruSAKiStGezahltEinzE0108003)"]', 'Kirchensteuer laut Lohnsteuerbescheinigung steuerpflichtige Person / Ehemann / Person A')
    pyautogui.press('enter')

nr22 = ws[f'W{row}'].value
nr23 = ws[f'X{row}'].value
if nr22 is not None or nr23 is not None:
    nr22i23URL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageVor/BeitraegeZurAltersvorsorge'
    driver.get(nr22i23URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="delete_btn_readMode_Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbLoeschenModal"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edit_btn_Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)"]'))).click()
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)_fields(eruVORAVorE2000401)"]', nr23)
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeZurAltersvorsorge(0)_BeitraegeZurAltersvorsorgeMZB(0)_fields(eruVORAVorE2000801)"]', nr22)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="UpdateMzbItem/Startseite[0]/VAnlageVor[0]/BeitraegeZurAltersvorsorge[0]/BeitraegeZurAltersvorsorgeMZB[0]"]'))).click()

nr25 = ws[f'y{row}'].value
nr26 = ws[f'z{row}'].value
if nr25 is not None:
    nr25i26URL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageVor/BeitraegeInlGesKrankenPflegevers'
    driver.get(nr25i26URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="delete_btn_Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbLoeschenModal"]'))).click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edit_btn_Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)"]'))).click()
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)_fields(eruVORBeitr_g_KV_PV_InlANE2001203)"]', nr25)
    if nr26 is None:
        nr26=0
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_BeitraegeInlGesKrankenPflegevers(0)_MZBBeitraegeInlGesKrankenPflegeversMZB(0)_fields(eruVORBeitr_g_KV_PV_InlANE2001505)"]', nr26)
    pyautogui.press('enter')

nr27 = ws[f'AS{row}'].value
if nr27 is not None:
    nr27URL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageVor/WeitereSonstigeVorsorgeaufwendungen'
    driver.get(nr27URL)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[4]/form/div[5]/main/div/div[4]/div/div/div[1]/div/div[2]/div[2]/div/div/div[3]/div/input[2]'))).click()
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.press('backspace')
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageVor(0)_WeitereSonstigeVorsorgeaufwendungen(0)_WeitSonsVorAW(0)_fields(eruVORWeit_Sons_VorAWPersE2004403)"]', nr27)
    pyautogui.press('enter')

#KRANKENGELD
krankengeld = ws[f'AU{row}'].value
if krankengeld is not None:
    krankengeldURL = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VHauptvordruck/Einkommensersatzleistungen'
    driver.get(krankengeldURL)
    try:
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="delete_btn_Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)"]'))).click()
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="confirm_mzbLoeschenModal"]'))).click()
    except:
        pass
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edit_btn_Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)"]'))).click()
    wait_and_send_keys('//*[@id="Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)_EinkErsInl(0)_fields(eruESt1AEink_ErsInlEinzE0104110)"]', 'Krankengeld')
    wait_and_send_keys('//*[@id="Startseite(0)_VHauptvordruck(0)_Einkommensersatzleistungen(0)_MZBEinkErs(0)_EinkErsInl(0)_fields(eruESt1AEink_ErsInlEinzE0104113)"]', krankengeld)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="CreateMzbItem/Startseite[0]/VHauptvordruck[0]/Einkommensersatzleistungen[0]/MZBEinkErs[0]/EinkErsInl[0]"]'))).click()


#ZAROBKI W POLSCE WA-EST

WaEstUrl = 'https://www.elster.de/eportal/interpreter/eingabe/est-2023/Startseite/VAnlageWAESt'
time.sleep(1)
driver.get(WaEstUrl)


ZarobkiMezaNiem = ws[f'U{row}'].value
ZarobkiMezaNiem = int(ZarobkiMezaNiem)
ZarobkiZonyNiem = ws[f'V{row}'].value
ZarobkiMezaNiem = int(ZarobkiMezaNiem)
ZarobkiZonyNiem = int(ZarobkiZonyNiem)
checkbox1 = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]')
checkbox2 = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]')
if ZarobkiMezaNiem + ZarobkiZonyNiem < 20000:
    if checkbox1.is_selected():
        if checkbox2.is_selected():
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]'))).click()
    else:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]'))).click()
else:
    if checkbox2.is_selected():
        if checkbox1.is_selected():
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0109906)"]'))).click()
    else:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStEheg_EU_EWRE0105901)"]'))).click()
try:
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105303)"]'))).clear()
except:
    pass
time.sleep(1)
input_field = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105301)"]')
input_field.clear()
time.sleep(.5)
wait_and_send_keys('//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105301)"]', ZarobkiMezaNiem)
if ZarobkiZonyNiem is not None:
    ZarobkiZonyNiem = int(ZarobkiZonyNiem)
    input_field = driver.find_element(By.XPATH, '//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105303)"]')
    input_field.clear()
    time.sleep(.5)
    wait_and_send_keys('//*[@id="Startseite(0)_VAnlageWAESt(0)_fields(eruWA_EStAntrag_unb_StpflE0105303)"]', ZarobkiZonyNiem)
time.sleep(1)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="SwitchModusPruefen"]'))).click()
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="steuerberechnungLink"]'))).click()
input("Wciśnij Enter, aby zakończyć działanie programu...")

